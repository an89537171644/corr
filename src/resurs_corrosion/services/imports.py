from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import Session

from ..domain import (
    ActionInput,
    AssetCreate,
    CheckType,
    ElementCreate,
    EnvironmentCategory,
    ImportFormat,
    ImportIssue,
    ImportSummary,
    InspectionCreate,
    MaterialInput,
    SectionDefinition,
    SectionType,
    ThicknessMeasurement,
    ZoneDefinition,
)
from ..storage import (
    create_asset,
    create_element,
    create_inspection,
    get_element_by_asset_and_code,
    get_inspection_by_element_and_code,
    update_element,
    update_inspection,
)


class ImportValidationError(ValueError):
    pass


@dataclass
class RowRecord:
    reference: str
    data: Dict[str, object]


@dataclass
class ImportAccumulator:
    dataset: str
    source_format: ImportFormat
    rows_processed: int = 0
    created_count: int = 0
    updated_count: int = 0
    warnings: List[str] = field(default_factory=list)
    errors: List[ImportIssue] = field(default_factory=list)

    def add_error(self, row_reference: str, message: str) -> None:
        self.errors.append(ImportIssue(row_reference=row_reference, message=message))

    def build(self) -> ImportSummary:
        return ImportSummary(
            dataset=self.dataset,
            source_format=self.source_format,
            rows_processed=self.rows_processed,
            created_count=self.created_count,
            updated_count=self.updated_count,
            warning_count=len(self.warnings),
            error_count=len(self.errors),
            warnings=self.warnings,
            errors=self.errors,
        )


def import_assets(session: Session, filename: str, payload: bytes) -> ImportSummary:
    source_format = detect_import_format(filename)
    rows = read_single_table(filename, payload)
    summary = ImportAccumulator(dataset="assets", source_format=source_format, rows_processed=len(rows))

    for row in rows:
        try:
            asset = AssetCreate(
                name=require_text(row.data, "name"),
                address=optional_text(row.data, "address"),
                commissioned_year=optional_int(row.data, "commissioned_year"),
                purpose=optional_text(row.data, "purpose"),
                responsibility_class=optional_text(row.data, "responsibility_class"),
            )
            create_asset(session, asset)
            summary.created_count += 1
        except Exception as exc:
            summary.add_error(row.reference, normalize_exception(exc))

    return summary.build()


def import_elements(session: Session, asset_id: int, filename: str, payload: bytes) -> ImportSummary:
    source_format = detect_import_format(filename)
    summary = ImportAccumulator(dataset="elements", source_format=source_format)

    if source_format == ImportFormat.CSV:
        grouped = group_element_csv_rows(read_single_table(filename, payload))
        summary.rows_processed = sum(len(item["rows"]) for item in grouped.values())
    else:
        grouped = group_element_xlsx_rows(payload)
        summary.rows_processed = sum(len(item["rows"]) for item in grouped.values())

    for element_code, item in grouped.items():
        reference = ",".join(item["rows"])
        try:
            payload_model = build_element_payload(item["element"], item["zones"])
            existing = get_element_by_asset_and_code(session, asset_id, element_code)
            if existing is None:
                create_element(session, asset_id, payload_model)
                summary.created_count += 1
            else:
                update_element(session, existing, payload_model)
                summary.updated_count += 1
        except Exception as exc:
            summary.add_error(reference, normalize_exception(exc))

    return summary.build()


def import_inspections(session: Session, element_id: int, filename: str, payload: bytes) -> ImportSummary:
    source_format = detect_import_format(filename)
    summary = ImportAccumulator(dataset="inspections", source_format=source_format)

    if source_format == ImportFormat.CSV:
        grouped = group_inspection_csv_rows(read_single_table(filename, payload))
        summary.rows_processed = sum(len(item["rows"]) for item in grouped.values())
    else:
        grouped = group_inspection_xlsx_rows(payload)
        summary.rows_processed = sum(len(item["rows"]) for item in grouped.values())

    for inspection_key, item in grouped.items():
        reference = ",".join(item["rows"])
        try:
            payload_model = build_inspection_payload(item["inspection"], item["measurements"])
            existing = None
            if payload_model.inspection_code:
                existing = get_inspection_by_element_and_code(session, element_id, payload_model.inspection_code)

            if existing is None:
                create_inspection(session, element_id, payload_model)
                summary.created_count += 1
            else:
                update_inspection(session, existing, payload_model)
                summary.updated_count += 1
        except Exception as exc:
            summary.add_error(reference or inspection_key, normalize_exception(exc))

    return summary.build()


def detect_import_format(filename: str) -> ImportFormat:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return ImportFormat.CSV
    if lowered.endswith(".xlsx"):
        return ImportFormat.XLSX
    raise ImportValidationError("Unsupported file format. Use .csv or .xlsx.")


def read_single_table(filename: str, payload: bytes) -> List[RowRecord]:
    source_format = detect_import_format(filename)
    if source_format == ImportFormat.CSV:
        return read_csv_rows(payload)

    workbook_rows = read_xlsx_tables(payload)
    if len(workbook_rows) != 1:
        raise ImportValidationError("XLSX import for this dataset expects exactly one worksheet.")
    return next(iter(workbook_rows.values()))


def read_csv_rows(payload: bytes) -> List[RowRecord]:
    text = decode_csv_bytes(payload)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames is None:
        raise ImportValidationError("CSV file does not contain a header row.")

    rows: List[RowRecord] = []
    for index, raw_row in enumerate(reader, start=2):
        cleaned = {normalize_field_name(key): clean_value(value) for key, value in raw_row.items() if key is not None}
        if any(value is not None for value in cleaned.values()):
            rows.append(RowRecord(reference=f"row {index}", data=cleaned))
    return rows


def decode_csv_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportValidationError("Unable to decode CSV file. Supported encodings: UTF-8 and CP1251.")


def read_xlsx_tables(payload: bytes) -> Dict[str, List[RowRecord]]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    tables: Dict[str, List[RowRecord]] = {}

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next((idx for idx, row in enumerate(rows) if any(cell is not None for cell in row)), None)
        if header_index is None:
            continue

        header = [normalize_field_name(cell) for cell in rows[header_index]]
        sheet_rows: List[RowRecord] = []
        for offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = list(row)
            cleaned = {
                header[col_index]: clean_value(values[col_index] if col_index < len(values) else None)
                for col_index in range(len(header))
                if header[col_index]
            }
            if any(value is not None for value in cleaned.values()):
                sheet_rows.append(RowRecord(reference=f"{sheet.title}!{offset}", data=cleaned))

        tables[normalize_field_name(sheet.title)] = sheet_rows

    return tables


def group_element_csv_rows(rows: List[RowRecord]) -> Dict[str, dict]:
    grouped: Dict[str, dict] = {}
    for row in rows:
        element_code = require_text(row.data, "element_id")
        zone = build_zone_dict_from_flat_row(row.data)
        bucket = grouped.setdefault(element_code, {"element": dict(row.data), "zones": [], "rows": []})
        merge_row_data(bucket["element"], row.data)
        bucket["zones"].append(zone)
        bucket["rows"].append(row.reference)
    return grouped


def group_element_xlsx_rows(payload: bytes) -> Dict[str, dict]:
    sheets = read_xlsx_tables(payload)
    if "elements" not in sheets or "zones" not in sheets:
        raise ImportValidationError("XLSX workbook for elements must contain 'elements' and 'zones' sheets.")

    grouped: Dict[str, dict] = {}
    for row in sheets["elements"]:
        element_code = require_text(row.data, "element_id")
        grouped[element_code] = {"element": dict(row.data), "zones": [], "rows": [row.reference]}

    for row in sheets["zones"]:
        element_code = require_text(row.data, "element_id")
        if element_code not in grouped:
            raise ImportValidationError(f"{row.reference}: zone references unknown element_id '{element_code}'.")
        grouped[element_code]["zones"].append(build_zone_dict(row.data))
        grouped[element_code]["rows"].append(row.reference)

    return grouped


def group_inspection_csv_rows(rows: List[RowRecord]) -> Dict[str, dict]:
    grouped: Dict[str, dict] = {}
    for row in rows:
        inspection_code = optional_text(row.data, "inspection_code")
        if inspection_code is None:
            performed_at = require_date(row.data, "performed_at")
            method = require_text(row.data, "method")
            inspection_key = f"{performed_at.isoformat()}::{method}"
        else:
            inspection_key = inspection_code

        bucket = grouped.setdefault(inspection_key, {"inspection": dict(row.data), "measurements": [], "rows": []})
        merge_row_data(bucket["inspection"], row.data)
        measurement = build_measurement_dict_from_flat_row(row.data)
        if measurement is not None:
            bucket["measurements"].append(measurement)
        bucket["rows"].append(row.reference)
    return grouped


def group_inspection_xlsx_rows(payload: bytes) -> Dict[str, dict]:
    sheets = read_xlsx_tables(payload)
    if "inspections" not in sheets:
        raise ImportValidationError("XLSX workbook for inspections must contain an 'inspections' sheet.")

    grouped: Dict[str, dict] = {}
    for row in sheets["inspections"]:
        inspection_key = optional_text(row.data, "inspection_code")
        if inspection_key is None:
            inspection_key = f"{require_date(row.data, 'performed_at').isoformat()}::{require_text(row.data, 'method')}"
        grouped[inspection_key] = {"inspection": dict(row.data), "measurements": [], "rows": [row.reference]}

    for row in sheets.get("measurements", []):
        inspection_code = optional_text(row.data, "inspection_code")
        if inspection_code is None:
            raise ImportValidationError(f"{row.reference}: measurements sheet requires inspection_code.")
        if inspection_code not in grouped:
            raise ImportValidationError(f"{row.reference}: measurement references unknown inspection_code '{inspection_code}'.")
        grouped[inspection_code]["measurements"].append(build_measurement_dict(row.data))
        grouped[inspection_code]["rows"].append(row.reference)

    return grouped


def build_element_payload(element_row: Dict[str, object], zones: List[Dict[str, object]]) -> ElementCreate:
    return ElementCreate(
        element_id=require_text(element_row, "element_id"),
        element_type=require_text(element_row, "element_type"),
        steel_grade=optional_text(element_row, "steel_grade"),
        work_scheme=optional_text(element_row, "work_scheme"),
        operating_zone=optional_text(element_row, "operating_zone"),
        environment_category=parse_environment_category(require_text(element_row, "environment_category")),
        current_service_life_years=optional_float(element_row, "current_service_life_years") or 0.0,
        section=build_section_definition(element_row),
        zones=[ZoneDefinition.model_validate(zone) for zone in zones],
        material=build_material_input(element_row),
        action=build_action_input(element_row),
    )


def build_inspection_payload(
    inspection_row: Dict[str, object],
    measurements: List[Dict[str, object]],
) -> InspectionCreate:
    return InspectionCreate(
        inspection_code=optional_text(inspection_row, "inspection_code"),
        performed_at=require_date(inspection_row, "performed_at"),
        method=require_text(inspection_row, "method"),
        executor=optional_text(inspection_row, "executor"),
        findings=optional_text(inspection_row, "findings"),
        measurements=[ThicknessMeasurement.model_validate(item) for item in measurements],
    )


def build_section_definition(row: Dict[str, object]) -> SectionDefinition:
    return SectionDefinition(
        section_type=parse_section_type(require_text(row, "section_type")),
        reference_thickness_mm=optional_float(row, "reference_thickness_mm"),
        width_mm=optional_float(row, "width_mm"),
        thickness_mm=optional_float(row, "thickness_mm"),
        height_mm=optional_float(row, "height_mm"),
        flange_width_mm=optional_float(row, "flange_width_mm"),
        web_thickness_mm=optional_float(row, "web_thickness_mm"),
        flange_thickness_mm=optional_float(row, "flange_thickness_mm"),
        area0_mm2=optional_float(row, "area0_mm2"),
        inertia0_mm4=optional_float(row, "inertia0_mm4"),
        section_modulus0_mm3=optional_float(row, "section_modulus0_mm3"),
    )


def build_material_input(row: Dict[str, object]) -> MaterialInput:
    return MaterialInput(
        fy_mpa=require_float(row, "fy_mpa"),
        gamma_m=optional_float(row, "gamma_m") or 1.0,
        stability_factor=optional_float(row, "stability_factor") or 1.0,
    )


def build_action_input(row: Dict[str, object]) -> ActionInput:
    return ActionInput(
        check_type=parse_check_type(require_text(row, "check_type")),
        demand_value=require_float(row, "demand_value"),
        demand_growth_factor_per_year=optional_float(row, "demand_growth_factor_per_year") or 0.0,
    )


def build_zone_dict_from_flat_row(row: Dict[str, object]) -> Dict[str, object]:
    zone_id = optional_text(row, "zone_id")
    if zone_id is None:
        raise ImportValidationError("Missing required field 'zone_id'.")
    return build_zone_dict(row)


def build_zone_dict(row: Dict[str, object]) -> Dict[str, object]:
    return {
        "zone_id": require_text(row, "zone_id"),
        "role": require_text(row, "role"),
        "initial_thickness_mm": require_float(row, "initial_thickness_mm"),
        "exposed_surfaces": optional_int(row, "exposed_surfaces") or 1,
        "pitting_factor": optional_float(row, "pitting_factor") or 0.0,
        "pit_loss_mm": optional_float(row, "pit_loss_mm") or 0.0,
    }


def build_measurement_dict_from_flat_row(row: Dict[str, object]) -> Optional[Dict[str, object]]:
    zone_id = optional_text(row, "zone_id")
    if zone_id is None:
        return None
    return build_measurement_dict(row)


def build_measurement_dict(row: Dict[str, object]) -> Dict[str, object]:
    return {
        "zone_id": require_text(row, "zone_id"),
        "point_id": optional_text(row, "point_id"),
        "thickness_mm": require_float(row, "thickness_mm"),
        "error_mm": optional_float(row, "error_mm") or 0.0,
        "measured_at": optional_date(row, "measured_at"),
        "quality": optional_float(row, "quality") or 1.0,
    }


def normalize_field_name(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_value(value: object) -> object:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def merge_row_data(target: Dict[str, object], source: Dict[str, object]) -> None:
    for key, value in source.items():
        if target.get(key) in (None, "") and value not in (None, ""):
            target[key] = value


def require_text(row: Dict[str, object], field_name: str) -> str:
    value = optional_text(row, field_name)
    if value is None:
        raise ImportValidationError(f"Missing required field '{field_name}'.")
    return value


def optional_text(row: Dict[str, object], field_name: str) -> Optional[str]:
    value = row.get(field_name)
    if value is None:
        return None
    return str(value).strip() or None


def require_float(row: Dict[str, object], field_name: str) -> float:
    value = optional_float(row, field_name)
    if value is None:
        raise ImportValidationError(f"Missing required numeric field '{field_name}'.")
    return value


def optional_float(row: Dict[str, object], field_name: str) -> Optional[float]:
    value = row.get(field_name)
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError as exc:
        raise ImportValidationError(f"Field '{field_name}' must be numeric.") from exc


def optional_int(row: Dict[str, object], field_name: str) -> Optional[int]:
    value = row.get(field_name)
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError as exc:
        raise ImportValidationError(f"Field '{field_name}' must be an integer.") from exc


def require_date(row: Dict[str, object], field_name: str) -> date:
    value = optional_date(row, field_name)
    if value is None:
        raise ImportValidationError(f"Missing required date field '{field_name}'.")
    return value


def optional_date(row: Dict[str, object], field_name: str) -> Optional[date]:
    value = row.get(field_name)
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ImportValidationError(f"Field '{field_name}' must be a valid date.")


def parse_environment_category(value: str) -> EnvironmentCategory:
    try:
        return EnvironmentCategory(value.upper())
    except ValueError as exc:
        raise ImportValidationError(f"Unsupported environment category '{value}'.") from exc


def parse_section_type(value: str) -> SectionType:
    normalized = value.strip().lower()
    try:
        return SectionType(normalized)
    except ValueError as exc:
        raise ImportValidationError(f"Unsupported section type '{value}'.") from exc


def parse_check_type(value: str) -> CheckType:
    normalized = value.strip().lower()
    try:
        return CheckType(normalized)
    except ValueError as exc:
        raise ImportValidationError(f"Unsupported check type '{value}'.") from exc


def normalize_exception(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        return "; ".join(error["msg"] for error in exc.errors())
    return str(exc)
